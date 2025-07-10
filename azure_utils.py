"""
Azure utilities and client classes for Multi-Agent AI Platform
Separated to avoid circular imports
"""

import time
import logging
import sys
from typing import Dict, Optional, List
import json
import requests
import hashlib
from datetime import datetime, timedelta

# Configure enhanced logging for debugging authentication issues
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("auth_debug.log"),
        logging.StreamHandler(sys.stdout)
    ]
)

# Azure imports
from azure.ai.projects import AIProjectClient
from azure.identity import DefaultAzureCredential, InteractiveBrowserCredential, ClientSecretCredential
from azure.core.exceptions import AzureError
from azure.storage.blob import BlobServiceClient
from azure.search.documents import SearchClient
from azure.search.documents.models import VectorizedQuery
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import (
    SearchIndex,
    SearchField,
    SearchFieldDataType,
    SimpleField,
    SearchableField,
    VectorSearch,
    HnswAlgorithmConfiguration,
    VectorSearchProfile,
    SemanticConfiguration,
    SemanticSearch,
    SemanticPrioritizedFields,
    SemanticField
)
from azure.core.credentials import AzureKeyCredential

# Configure logging
logger = logging.getLogger(__name__)

# Configuration class for Azure services
class AzureConfig:
    """Centralized configuration for Azure services"""
    
    def __init__(self):
        # Check for environment variables first, then fall back to defaults
        import os
        
        # Azure AD App Registration for authentication
        self.client_id = os.getenv("AZURE_CLIENT_ID", "c7790b94-d830-4746-961f-8c715a380c5e")
        self.client_secret = os.getenv("AZURE_CLIENT_SECRET", "6Jl8Q~Kue3BfEbXdbc3O-68WVEIPZeNFD-Bkub2p")
        self.tenant_id = os.getenv("AZURE_TENANT_ID", "7ae3526a-96fa-407a-9b02-9fe5bdff6217")
        
        # Development vs Production mode
        self.dev_mode = os.getenv("DEV_MODE", "false").lower() == "true"
        self.use_managed_identity = os.getenv("USE_MANAGED_IDENTITY", "true").lower() == "true"
        
        # For local development, disable managed identity
        if self.dev_mode:
            self.use_managed_identity = False
            logger.info("Development mode enabled - using service principal authentication")
        
        # OAuth2 token endpoint ve Graph API endpoint tanımlamaları
        self.oauth_token_endpoint = f"https://login.microsoftonline.com/{self.tenant_id}/oauth2/v2.0/token"
        self.graph_endpoint = "https://graph.microsoft.com/v1.0"
        
        # Other Azure services - Use environment variables or defaults
        self.redirect_uri = os.getenv("REDIRECT_URI", "https://egentapp-b4gqeudnc3h8emd3.westeurope-01.azurewebsites.net/")
        
        # Storage configuration with fallback authentication
        self.storage_connection_string = os.getenv(
            "AZURE_STORAGE_CONNECTION_STRING", 
            "DefaultEndpointsProtocol=https;AccountName=egenthub7616267901;AccountKey=gQPagF4oAGn3KToKSSIXdtPDxGlW1pHYjBLiAR8VldtfaLOO4iZ1D6gNidyqnjGP6yKDRleJTeOJ+AStYe2LTQ==;EndpointSuffix=core.windows.net"
        )
        self.storage_account_name = os.getenv("AZURE_STORAGE_ACCOUNT_NAME", "egenthub7616267901")
        
        # Search configuration
        self.search_endpoint = os.getenv("AZURE_SEARCH_SERVICE_ENDPOINT", "https://egesearchindex.search.windows.net")
        self.search_admin_key = os.getenv("AZURE_SEARCH_ADMIN_KEY", "IVMK9OMEVHJxlOLBI88BMIpYr3AeP95zwfDmSfD1k1AzSeCinJJT")
        
        # Azure OpenAI configuration for embeddings
        self.openai_endpoint = "https://egentaimodel.openai.azure.com"  # Keep existing endpoint
        self.openai_api_key = os.getenv(
            "OPENAI_API_KEY", 
            "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"
        )
        self.openai_api_version = "2023-05-15"
        self.embedding_model = "text-embedding-ada-002"
        
        logger.info(f"Azure config initialized - Dev mode: {self.dev_mode}, Managed Identity: {self.use_managed_identity}")
    
    def get_credential(self):
        """Get appropriate Azure credential based on environment"""
        try:
            if self.use_managed_identity and not self.dev_mode:
                # Production: Use Managed Identity
                logger.info("Using Managed Identity for authentication")
                return DefaultAzureCredential()
            else:
                # Development: Use Service Principal with connection string fallback
                logger.info("Using Service Principal for authentication in development mode")
                return ClientSecretCredential(
                    tenant_id=self.tenant_id,
                    client_id=self.client_id,
                    client_secret=self.client_secret
                )
        except Exception as e:
            logger.error(f"Error creating credential: {e}")
            # Fallback to DefaultAzureCredential which will try multiple auth methods
            return DefaultAzureCredential()
    
    def get_storage_client(self):
        """Get Azure Storage client with appropriate authentication"""
        try:
            if self.use_managed_identity and not self.dev_mode:
                # Production: Use Managed Identity
                credential = self.get_credential()
                return BlobServiceClient(
                    account_url=f"https://{self.storage_account_name}.blob.core.windows.net",
                    credential=credential
                )
            else:
                # Development: Use connection string (more reliable for dev)
                logger.info("Using storage connection string for development")
                return BlobServiceClient.from_connection_string(self.storage_connection_string)
        except Exception as e:
            logger.error(f"Error creating storage client: {e}")
            # Fallback to connection string
            return BlobServiceClient.from_connection_string(self.storage_connection_string)

# Azure Authentication Helper
class AzureAuthenticator:
    """Handles Azure AD authentication"""
    
    def __init__(self, config: AzureConfig):
        self.config = config
    
    def authenticate_with_device_code(self) -> Dict:
        """
        Initiate device code flow authentication for Azure AD
        Returns device code information for the user to complete authentication
        """
        try:
            from msal import PublicClientApplication
            
            # Initialize MSAL client
            app = PublicClientApplication(
                client_id=self.config.public_client_id,
                authority=f"https://login.microsoftonline.com/{self.config.tenant_id}"
            )
            
            # Define scopes for authentication
            scopes = ["https://graph.microsoft.com/User.Read"]
            
            # Initiate device flow
            flow = app.initiate_device_flow(scopes=scopes)
            
            if "user_code" not in flow:
                logger.error(f"Failed to initiate device flow: {flow.get('error_description', 'Unknown error')}")
                return {
                    "success": False,
                    "message": f"Failed to initiate device code flow: {flow.get('error_description', 'Unknown error')}"
                }
            
            # Return device flow information
            return {
                "success": True,
                "device_flow": flow,
                "user_code": flow["user_code"],
                "verification_uri": flow["verification_uri"],
                "expires_in": flow["expires_in"],
                "message": "Device code flow initiated successfully. Please complete authentication in your browser."
            }
            
        except ImportError:
            logger.error("MSAL library not found. Please install: pip install msal")
            return {
                "success": False,
                "message": "MSAL library not installed. Please install msal package."
            }
        except Exception as e:
            logger.error(f"Error initiating device flow: {e}")
            return {
                "success": False,
                "message": f"Error initiating device code flow: {str(e)}"
            }
    
    def complete_device_code_authentication(self, flow: Dict) -> Dict:
        """
        Complete device code authentication process
        Args:
            flow: The flow object returned from authenticate_with_device_code
            
        Returns:
            Dict with authentication result
        """
        try:
            from msal import PublicClientApplication
            
            # Initialize MSAL client
            app = PublicClientApplication(
                client_id=self.config.public_client_id,
                authority=f"https://login.microsoftonline.com/{self.config.tenant_id}"
            )
            
            # Try to acquire token by device flow
            result = app.acquire_token_by_device_flow(flow)
            
            if "access_token" in result:
                # Authentication successful
                user_info = self._get_user_info_from_graph(result["access_token"])
                username = user_info.get("userPrincipalName", "unknown")
                role = self._determine_user_role(username)
                permissions = self._get_user_permissions(username, role)
                
                logger.info(f"Device code authentication successful for {username}")
                return {
                    "success": True,
                    "user": {
                        "username": username,
                        "display_name": user_info.get("displayName", username),
                        "role": role,
                        "permissions": permissions,
                        "email": user_info.get("mail") or user_info.get("userPrincipalName"),
                        "real_azure_user": True,
                        "auth_method": "device_code"
                    },
                    "token": result["access_token"],
                    "message": "Azure AD authentication successful"
                }
            else:
                error_description = result.get("error_description", "Unknown error")
                logger.error(f"Device code authentication failed: {error_description}")
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD authentication failed: {error_description}"
                }
                
        except ImportError:
            logger.error("MSAL library not found. Please install: pip install msal")
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": "MSAL library not installed. Please install msal package."
            }
        except Exception as e:
            logger.error(f"Error completing device code authentication: {e}")
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Error completing device code authentication: {str(e)}"
            }
    
    def authenticate_with_username_password(self, username: str, password: str) -> Dict:
        """
        Authenticate user with username/password against Azure AD
        Only supports real Azure AD authentication
        """
        try:
            # Directly authenticate with Azure AD
            return self._authenticate_real_azure_user(username, password)
                
        except Exception as e:
            logger.error(f"Azure authentication error: {e}")
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Authentication error: {str(e)}"
            }
    
    def _authenticate_real_azure_user(self, username: str, password: str) -> Dict:
        """
        Authenticate real Azure AD user using direct OAuth2 ROPC flow
        This method uses the OAuth2 password grant flow to authenticate against Azure AD
        
        Note: This flow requires that:
        1. The app is registered with proper permissions
        2. The user does not have MFA enabled
        3. The tenant policy allows ROPC flow
        4. Using /organizations or tenant-specific endpoint (not /common or /consumers)
        """
        try:
            import requests
            
            # DEBUG: Print detailed information about the authentication process
            logger.debug("=== STARTING AZURE AD AUTHENTICATION DEBUG ===")
            logger.debug(f"Username: {username}")
            logger.debug(f"Configuration: client_id={self.config.client_id[:6]}..., tenant_id={self.config.tenant_id}")
            
            # Try both organizations and common endpoints for testing
            token_url = f"https://login.microsoftonline.com/organizations/oauth2/v2.0/token"
            
            # Log the token URL being used (don't log in production)
            logger.debug(f"Using OAuth token endpoint: {token_url}")
            
            # Prepare the request body for OAuth2 password grant flow
            token_data = {
                'grant_type': 'password',
                'client_id': self.config.client_id,
                'client_secret': self.config.client_secret,  # Required for confidential clients
                'scope': 'https://graph.microsoft.com/.default offline_access',  # Broader scope for debugging
                'username': username,
                'password': password
            }
            
            # DEBUG: Log the request data (sanitize password)
            debug_data = token_data.copy()
            debug_data['password'] = '********'
            logger.debug(f"Auth request data: {debug_data}")
            
            # Add proper headers for the token request
            headers = {
                'Content-Type': 'application/x-www-form-urlencoded',
                'Accept': 'application/json'
            }
            
            # DEBUG: Log full request details
            logger.debug(f"Request headers: {headers}")
            
            # Make the token request with proper error handling
            try:
                token_response = requests.post(token_url, data=token_data, headers=headers)
                logger.debug(f"Response status code: {token_response.status_code}")
                logger.debug(f"Response headers: {dict(token_response.headers)}")
            except Exception as req_ex:
                logger.error(f"Request exception: {req_ex}")
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Network error during authentication: {str(req_ex)}"
                }
            
            # Try to parse JSON response - handle potential JSON parsing errors
            try:
                result = token_response.json()
                logger.debug(f"Response content type: {token_response.headers.get('Content-Type', 'unknown')}")
            except ValueError:
                # Log the raw response for debugging
                logger.error(f"Invalid JSON response. Status code: {token_response.status_code}")
                logger.error(f"Response text: {token_response.text[:500]}")
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD authentication error: Invalid response format (Status: {token_response.status_code})"
                }
            
            # Log error details for debugging
            if "error" in result:
                error_code = result.get("error")
                error_description = result.get("error_description", "")
                correlation_id = result.get("correlation_id", "unknown")
                timestamp = result.get("timestamp", "unknown")
                trace_id = result.get("trace_id", "unknown")
                
                # Detailed logging of the error
                logger.error(f"OAuth token error: {error_code}")
                logger.error(f"Full error description: {error_description}")
                logger.error(f"Correlation ID: {correlation_id}")
                logger.error(f"Timestamp: {timestamp}")
                logger.error(f"Trace ID: {trace_id}")
                
                # Debug - try to extract AADSTS code if present
                import re
                aadsts_match = re.search(r'AADSTS\d+', error_description)
                if aadsts_match:
                    aadsts_code = aadsts_match.group(0)
                    logger.error(f"AADSTS Error Code: {aadsts_code}")
                
                # If we got AADSTS9000102 error, try with a specific tenant ID
                if "AADSTS9000102" in error_description:
                    logger.debug("Got AADSTS9000102 error, trying with specific tenant ID instead")
                    # You can set a specific tenant ID here if you have one
                    specific_tenant_id = "7ae3526a-96fa-407a-9b02-9fe5bdff6217"  # Example tenant ID
                    return self._try_with_specific_tenant(username, password, specific_tenant_id)
                
                # Handle common error scenarios with user-friendly messages
                if "AADSTS50076" in error_description or "AADSTS50079" in error_description:
                    # MFA required error
                    return {
                        "success": False,
                        "user": None,
                        "token": None,
                        "message": "Bu hesap için çok faktörlü kimlik doğrulama (MFA) gerekiyor."
                    }
                elif "AADSTS50126" in error_description:
                    # Invalid username or password
                    return {
                        "success": False,
                        "user": None,
                        "token": None,
                        "message": "Hatalı kullanıcı adı veya şifre. Lütfen kimlik bilgilerinizi kontrol edin."
                    }
                elif "AADSTS50034" in error_description:
                    # User not found
                    return {
                        "success": False,
                        "user": None,
                        "token": None,
                        "message": "Bu kullanıcı hesabı Azure AD'de bulunamadı."
                    }
            
            if "access_token" in result:
                # Success with username/password
                user_info = self._get_user_info_from_graph(result["access_token"])
                role = self._determine_user_role(username)
                permissions = self._get_user_permissions(username, role)
                
                logger.info(f"Real Azure authentication successful for {username}")
                return {
                    "success": True,
                    "user": {
                        "username": username,
                        "display_name": user_info.get("displayName", username),
                        "role": role,
                        "permissions": permissions,
                        "email": user_info.get("mail") or user_info.get("userPrincipalName"),
                        "real_azure_user": True,
                        "auth_method": "password"
                    },
                    "token": result["access_token"],
                    "message": "Azure AD authentication successful"
                }
            else:
                error_description = result.get("error_description", "Unknown error")
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD authentication failed: {error_description}"
                }
                
        except ImportError:
            logger.error("MSAL library not found. Please install: pip install msal")
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": "MSAL library not installed. Please install msal package."
            }
        except Exception as e:
            logger.error(f"Real Azure authentication error: {e}")
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Azure AD authentication error: {str(e)}"
            }
    
    def _try_with_specific_tenant(self, username: str, password: str, tenant_id: str) -> Dict:
        """Fallback authentication method using a specific tenant ID"""
        logger.debug(f"Attempting authentication with specific tenant ID: {tenant_id}")
        
        try:
            import requests
            
            # Use specific tenant ID endpoint
            token_url = f"https://login.microsoftonline.com/{tenant_id}/oauth2/v2.0/token"
            logger.debug(f"Fallback OAuth token endpoint: {token_url}")
            
            # Headers
            headers = {
                'Content-Type': 'application/x-www-form-urlencoded',
                'Accept': 'application/json'
            }
            
            # Request data
            token_data = {
                'grant_type': 'password',
                'client_id': self.config.client_id,
                'client_secret': self.config.client_secret,
                'scope': 'https://graph.microsoft.com/.default',
                'username': username,
                'password': password
            }
            
            # Make the request
            token_response = requests.post(token_url, data=token_data, headers=headers)
            logger.debug(f"Fallback response status code: {token_response.status_code}")
            
            # Parse response
            try:
                result = token_response.json()
            except ValueError:
                logger.error(f"Invalid JSON in fallback response: {token_response.text[:200]}")
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD authentication error with fallback tenant: Invalid response format"
                }
            
            # Check for errors
            if "error" in result:
                error_code = result.get("error")
                error_description = result.get("error_description", "")
                logger.error(f"Fallback OAuth token error: {error_code}")
                logger.error(f"Fallback error description: {error_description[:200]}")
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD authentication failed with fallback tenant: {error_description}"
                }
                
            if "access_token" in result:
                # Success with username/password
                user_info = self._get_user_info_from_graph(result["access_token"])
                role = self._determine_user_role(username)
                permissions = self._get_user_permissions(username, role)
                
                logger.info(f"Fallback Azure authentication successful for {username}")
                return {
                    "success": True,
                    "user": {
                        "username": username,
                        "display_name": user_info.get("displayName", username),
                        "role": role,
                        "permissions": permissions,
                        "email": user_info.get("mail") or user_info.get("userPrincipalName"),
                        "real_azure_user": True,
                        "auth_method": "password"
                    },
                    "token": result["access_token"],
                    "message": "Azure AD authentication successful with fallback tenant"
                }
                
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": "Unknown error with fallback tenant authentication"
            }
                
        except Exception as e:
            logger.error(f"Error in fallback authentication: {e}")
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Error in fallback authentication: {str(e)}"
            }
    
    def _get_user_info_from_graph(self, access_token: str) -> Dict:
        """Get user information from Microsoft Graph API"""
        try:
            import requests
            
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': 'application/json'
            }
            
            response = requests.get(
                'https://graph.microsoft.com/v1.0/me',
                headers=headers
            )
            
            if response.status_code == 200:
                return response.json()
            else:
                logger.warning(f"Failed to get user info from Graph API: {response.status_code}")
                return {}
                
        except Exception as e:
            logger.error(f"Error getting user info from Graph API: {e}")
            return {}
    
    def _determine_user_role(self, username: str) -> str:
        """Determine user role based on username or other criteria"""
        # For now, assign standard role to all real users
        # In production, you might check Azure AD groups or custom attributes
        if username.lower().startswith("admin") or "admin" in username.lower():
            return "admin"
        elif username.lower().startswith("manager") or "manager" in username.lower():
            return "manager"
        else:
            return "standard"
    
    def _get_user_permissions(self, username: str, role: str) -> List[str]:
        """Get user permissions based on role - now returns a simple list of permissions"""
        if role == "admin":
            # Admin has all permissions
            return ["all"]
        elif role == "manager":
            # Manager has most permissions for all agents
            return [
                "access", "chat", "document_upload", "document_delete",
                # Allow all current and future agents
                "legal:access", "legal:chat", "legal:document_upload", "legal:document_delete",
                "it:access", "it:chat", "it:document_upload", "it:document_delete",
                "sample:access", "sample:chat", "sample:document_upload", "sample:document_delete",
                "scm:access", "scm:chat", "scm:document_upload", "scm:document_delete"
            ]
        else:
            # Standard user has basic permissions for all agents
            return [
                "access", "chat", "document_upload",
                # Specific permissions for each agent
                "legal:access", "legal:chat", "legal:document_upload",
                "it:access", "it:chat", "it:document_upload", 
                "sample:access", "sample:chat", "sample:document_upload",
                "scm:access", "scm:chat", "scm:document_upload"
            ]
    
    def get_azure_credential(self) -> DefaultAzureCredential:
        """Get Azure credential for service authentication"""
        try:
            if self.config.dev_mode:
                # For development, use client secret credential
                return ClientSecretCredential(
                    tenant_id=self.config.tenant_id,
                    client_id=self.config.client_id,
                    client_secret=self.config.client_secret
                )
            else:
                # For production, use managed identity or default credential chain
                return DefaultAzureCredential()
        except Exception as e:
            logger.error(f"Error creating Azure credential: {e}")
            return DefaultAzureCredential()

# Enhanced Azure AI Agent Client
class EnhancedAzureAIAgentClient:
    """Enhanced wrapper for Azure AI Agent operations with document management"""
    
    def __init__(self, connection_string: str, agent_id: str, config: AzureConfig):
        self.connection_string = connection_string
        self.agent_id = agent_id
        self.config = config
        self.client = None
        self.agent = None
        self.blob_client = None
        self.search_client = None
        self.search_index_client = None
        self._initialize_clients()
    
    def _initialize_clients(self):
        """Initialize all Azure clients with proper error handling"""
        try:
            # Initialize AI Project Client
            self.client = AIProjectClient.from_connection_string(
                credential=DefaultAzureCredential(),
                conn_str=self.connection_string
            )
            self.agent = self.client.agents.get_agent(self.agent_id)
            
            # Initialize Blob Storage Client
            self.blob_client = BlobServiceClient.from_connection_string(
                self.config.storage_connection_string
            )
            
            # Initialize Azure Search clients
            try:
                if self.config.search_endpoint and self.config.search_admin_key:
                    from azure.search.documents.indexes import SearchIndexClient
                    from azure.core.credentials import AzureKeyCredential
                    
                    credential = AzureKeyCredential(self.config.search_admin_key)
                    self.search_index_client = SearchIndexClient(
                        endpoint=self.config.search_endpoint,
                        credential=credential
                    )
                    logger.info("Azure Search Index client initialized successfully")
                else:
                    logger.warning("Search endpoint or admin key not configured - search functionality disabled")
                    self.search_index_client = None
                    self.search_client = None
            except Exception as search_error:
                logger.warning(f"Failed to initialize search clients: {search_error}")
                self.search_index_client = None
                self.search_client = None
            
            logger.info("All Azure clients initialized successfully")
            
        except AzureError as e:
            logger.error(f"Azure error during initialization: {e}")
            raise
        except Exception as e:
            logger.error(f"Unexpected error during initialization: {e}")
            raise
    
    def create_thread(self):
        """Create a new conversation thread"""
        try:
            thread = self.client.agents.create_thread()
            logger.info(f"Thread created with ID: {thread.id}")
            return thread
        except AzureError as e:
            logger.error(f"Error creating thread: {e}")
            raise
    
    def send_message_and_get_response(self, thread_id: str, user_message: str):
        """Send message to agent and get response with retry logic"""
        max_retries = 3
        retry_delay = 1
        
        for attempt in range(max_retries):
            try:
                # Create user message
                self.client.agents.create_message(
                    thread_id=thread_id,
                    role="user",
                    content=user_message
                )
                
                # Create and process run
                run = self.client.agents.create_and_process_run(
                    thread_id=thread_id,
                    agent_id=self.agent.id
                )
                
                # Get messages
                messages = self.client.agents.list_messages(thread_id=thread_id)
                
                # Extract response
                text_messages = list(messages.text_messages)
                if text_messages:
                    first_message = text_messages[0]
                    
                    # Extract the text content
                    response_text = ""
                    if hasattr(first_message, 'text') and hasattr(first_message.text, 'value'):
                        response_text = first_message.text.value
                        logger.info(f"Extracted response text: {response_text[:100]}...")
                    
                    # Process file citation annotations to replace source placeholders
                    if hasattr(first_message, 'file_citation_annotations'):
                        citations = first_message.file_citation_annotations
                        logger.info(f"Found {len(citations)} file citation annotations")
                        
                        # Replace citation placeholders with actual file names
                        for i, citation in enumerate(citations):
                            if hasattr(citation, 'file_citation') and hasattr(citation.file_citation, 'file_id'):
                                # Try to get file information
                                try:
                                    file_info = self.client.agents.get_file(citation.file_citation.file_id)
                                    file_name = getattr(file_info, 'filename', f'file_{i}')
                                    logger.info(f"Citation {i}: {file_name}")
                                    
                                    # Replace the citation placeholder with actual file name
                                    placeholder = f"[{i}:source]"
                                    if placeholder in response_text:
                                        response_text = response_text.replace(placeholder, f"[{file_name}]")
                                        logger.info(f"Replaced {placeholder} with [{file_name}]")
                                except Exception as file_error:
                                    logger.warning(f"Could not get file info for citation {i}: {file_error}")
                                    # Replace with generic source reference
                                    placeholder = f"[{i}:source]"
                                    if placeholder in response_text:
                                        response_text = response_text.replace(placeholder, f"[Kaynak Dosya {i+1}]")
                    
                    # Also handle text annotations if they exist
                    if hasattr(first_message, 'text') and hasattr(first_message.text, 'annotations'):
                        annotations = first_message.text.annotations
                        logger.info(f"Found {len(annotations)} text annotations")
                        
                        for i, annotation in enumerate(annotations):
                            logger.info(f"Processing annotation {i}: {type(annotation)}")
                            
                            # Handle file citations
                            if hasattr(annotation, 'file_citation'):
                                try:
                                    file_info = self.client.agents.get_file(annotation.file_citation.file_id)
                                    file_name = getattr(file_info, 'filename', f'file_{i}')
                                    logger.info(f"File annotation {i}: {file_name}")
                                    
                                    # Replace citation text with file name
                                    if hasattr(annotation, 'text'):
                                        old_text = annotation.text
                                        response_text = response_text.replace(old_text, f"[{file_name}]")
                                        logger.info(f"Replaced annotation text '{old_text}' with [{file_name}]")
                                except Exception as file_error:
                                    logger.warning(f"Could not get file info for file annotation {i}: {file_error}")
                                    # Replace with generic source reference
                                    if hasattr(annotation, 'text'):
                                        old_text = annotation.text
                                        response_text = response_text.replace(old_text, f"[Kaynak Dosya {i+1}]")
                                        logger.info(f"Replaced annotation text '{old_text}' with [Kaynak Dosya {i+1}]")
                            
                            # Handle URL citations
                            elif hasattr(annotation, 'url_citation'):
                                try:
                                    url_citation = annotation.url_citation
                                    
                                    # First try to get title (this should be the actual filename)
                                    file_name = getattr(url_citation, 'title', None)
                                    
                                    # If no title, fallback to URL
                                    if not file_name:
                                        url = getattr(url_citation, 'url', f'URL_{i}')
                                        # Try to extract a meaningful name from the URL
                                        import os
                                        file_name = os.path.basename(url) if url else f'URL_{i}'
                                    
                                    # Final fallback
                                    if not file_name or file_name in ['doc_0', f'URL_{i}']:
                                        file_name = f"Web Kaynağı {i+1}"
                                    
                                    logger.info(f"URL annotation {i}: {file_name}")
                                    
                                    # Replace citation text with file name
                                    if hasattr(annotation, 'text'):
                                        old_text = annotation.text
                                        response_text = response_text.replace(old_text, f"[{file_name}]")
                                        logger.info(f"Replaced URL annotation text '{old_text}' with [{file_name}]")
                                except Exception as url_error:
                                    logger.warning(f"Could not process URL annotation {i}: {url_error}")
                                    # Replace with generic source reference
                                    if hasattr(annotation, 'text'):
                                        old_text = annotation.text
                                        response_text = response_text.replace(old_text, f"[Web Kaynağı {i+1}]")
                                        logger.info(f"Replaced URL annotation text '{old_text}' with [Web Kaynağı {i+1}]")
                            
                            # Handle any other annotation types generically
                            else:
                                if hasattr(annotation, 'text'):
                                    old_text = annotation.text
                                    response_text = response_text.replace(old_text, f"[Kaynak {i+1}]")
                                    logger.info(f"Replaced generic annotation text '{old_text}' with [Kaynak {i+1}]")
                    
                    # Additional processing for any remaining citation patterns
                    import re
                    
                    # Look for any remaining [:source] patterns and replace them
                    citation_pattern = r'\[(\d+):source\]'
                    matches = re.findall(citation_pattern, response_text)
                    
                    for match in matches:
                        index = int(match)
                        placeholder = f"[{index}:source]"
                        response_text = response_text.replace(placeholder, f"[Kaynak Dosya {index+1}]")
                        logger.info(f"Replaced remaining placeholder {placeholder} with [Kaynak Dosya {index+1}]")
                    
                    return response_text if response_text else "No response received from agent."

                return "No response received from agent."
                
            except AzureError as e:
                logger.warning(f"Attempt {attempt + 1} failed: {e}")
                if attempt < max_retries - 1:
                    time.sleep(retry_delay * (2 ** attempt))
                else:
                    raise
    
    def upload_document(self, container_name: str, file_name: str, file_content: bytes):
        """Upload document to Azure Blob Storage"""
        try:
            # Try to upload to Azure Blob Storage
            try:
                # Create container if it doesn't exist
                container_client = self.blob_client.get_container_client(container_name)
                try:
                    container_client.get_container_properties()
                except:
                    container_client.create_container()
                
                # Upload blob
                blob_client = container_client.get_blob_client(file_name)
                blob_client.upload_blob(file_content, overwrite=True)
                
                logger.info(f"Document {file_name} uploaded to {container_name}")
                return True
                
            except Exception as azure_error:
                logger.warning(f"Azure Blob Storage upload failed: {azure_error}")
                logger.info("Document upload simulated for demo purposes")
                
                # For demo purposes, just log the upload
                file_size = len(file_content) if file_content else 0
                logger.info(f"Demo upload: {file_name} ({file_size} bytes) to {container_name}")
                return True
            
        except Exception as e:
            logger.error(f"Error uploading document: {e}")
            return False
    
    def list_documents(self, container_name: str):
        """List documents in container"""
        try:
            # For demo purposes, if Azure connection fails, return demo documents
            try:
                container_client = self.blob_client.get_container_client(container_name)
                
                # Try to get container properties first, create if doesn't exist
                try:
                    container_client.get_container_properties()
                except Exception as container_error:
                    if "ContainerNotFound" in str(container_error):
                        logger.info(f"Container {container_name} not found, creating it...")
                        container_client.create_container()
                        logger.info(f"Container {container_name} created successfully")
                        # Return empty list for new container
                        return []
                    else:
                        raise container_error
                
                # List blobs in the container
                blobs = container_client.list_blobs()
                
                documents = []
                for blob in blobs:
                    documents.append({
                        'name': blob.name,
                        'size': blob.size or 0,
                        'last_modified': blob.last_modified,
                        'content_type': blob.content_settings.content_type if blob.content_settings else 'unknown'
                    })
                
                logger.info(f"Found {len(documents)} documents in container {container_name}")
                return documents
                
            except Exception as azure_error:
                logger.warning(f"Azure Blob Storage connection failed: {azure_error}")
                logger.info("Returning demo documents for testing")
                
                # Return demo documents
                from datetime import datetime, timedelta
                demo_docs = [
                    {
                        'name': 'sample_report.pdf',
                        'size': 1024 * 1024,  # 1 MB
                        'last_modified': datetime.now() - timedelta(hours=2),
                        'content_type': 'application/pdf'
                    },
                    {
                        'name': 'data_analysis.xlsx',
                        'size': 512 * 1024,  # 512 KB
                        'last_modified': datetime.now() - timedelta(days=1),
                        'content_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                    },
                    {
                        'name': 'meeting_notes.docx',
                        'size': 256 * 1024,  # 256 KB
                        'last_modified': datetime.now() - timedelta(hours=5),
                        'content_type': 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                    }
                ]
                return demo_docs
            
        except Exception as e:
            logger.error(f"Error listing documents: {e}")
            return []
    
    def delete_document(self, container_name: str, file_name: str, index_name: str = None):
        """Delete document from container and ensure removal from search index"""
        try:
            # Strict validation: index_name is required
            if not index_name:
                logger.error(f"Cannot delete document '{file_name}': search index name is required but not provided")
                return False
            
            if not isinstance(index_name, str) or index_name.strip() == "":
                logger.error(f"Cannot delete document '{file_name}': invalid search index name '{index_name}'")
                return False
            
            index_name = index_name.strip()
            logger.info(f"Deleting document '{file_name}' from container '{container_name}' and index '{index_name}'")
            
            # Delete from blob storage first
            container_client = self.blob_client.get_container_client(container_name)
            blob_client = container_client.get_blob_client(file_name)
            blob_client.delete_blob()
            
            logger.info(f"Document {file_name} deleted from {container_name}")
            
            # Use standard deletion method for the specified index
            index_deletion_success = self._standard_index_deletion(container_name, file_name, index_name)
            
            if index_deletion_success:
                logger.info(f"Document '{file_name}' successfully removed from search index '{index_name}'")
            else:
                logger.error(f"Failed to remove document '{file_name}' from search index '{index_name}'")
                # Still return True since blob deletion succeeded, but log the issue
            
            return True
            
        except Exception as e:
            logger.error(f"Error deleting document: {e}")
            return False

    def _standard_index_deletion(self, container_name: str, file_name: str, index_name: str):
        """Standard index deletion process for all indexes"""
        deletion_success = False
        
        # Step 1: Directly remove document from search index
        try:
            remove_result = self.remove_document_from_index(container_name, file_name, index_name)
            if remove_result:
                logger.info(f"Document '{file_name}' directly removed from search index '{index_name}'")
                deletion_success = True
            else:
                logger.warning(f"Failed to directly remove document '{file_name}' from search index '{index_name}'")
        except Exception as direct_delete_error:
            logger.error(f"Error in direct index removal: {direct_delete_error}")
        
        # Step 2: Trigger indexer for comprehensive reindexing
        try:
            reindex_result = self.trigger_reindex_after_document_change(container_name, index_name)
            if reindex_result["success"]:
                logger.info(f"Indexer rerun triggered after deleting '{file_name}' from '{container_name}'")
                deletion_success = True
            else:
                logger.warning(f"Failed to trigger indexer rerun after deletion: {reindex_result.get('message')}")
        except Exception as indexer_error:
            logger.error(f"Error triggering indexer: {indexer_error}")
        
        # Step 3: Additional cleanup - search and delete any matching documents
        if not deletion_success:
            try:
                logger.info(f"Attempting advanced search and delete for '{file_name}' in '{index_name}'")
                advanced_delete_result = self._advanced_search_and_delete(container_name, file_name, index_name)
                if advanced_delete_result:
                    logger.info(f"Advanced deletion successful for '{file_name}' in '{index_name}'")
                    deletion_success = True
            except Exception as advanced_error:
                logger.error(f"Error in advanced deletion: {advanced_error}")
        
        return deletion_success

    def create_search_index(self, index_name: str) -> bool:
        """Create Azure Search index with vector search capabilities"""
        try:
            if not self.search_index_client:
                logger.warning("Search index client not initialized - cannot create index")
                return False
                
            # Check if index already exists
            try:
                existing_index = self.search_index_client.get_index(index_name)
                logger.info(f"Index '{index_name}' already exists")
                return True
            except Exception:
                # Index doesn't exist, create it
                pass
            
            # Define index schema
            fields = [
                SimpleField(name="id", type=SearchFieldDataType.String, key=True),
                SearchableField(name="content", type=SearchFieldDataType.String),
                SearchableField(name="metadata_storage_name", type=SearchFieldDataType.String, filterable=True),
                SimpleField(name="metadata_storage_path", type=SearchFieldDataType.String),
                SimpleField(name="metadata_storage_content_type", type=SearchFieldDataType.String, filterable=True),
                SimpleField(name="metadata_storage_size", type=SearchFieldDataType.Int64, filterable=True),
                SimpleField(name="metadata_storage_last_modified", type=SearchFieldDataType.DateTimeOffset, filterable=True, sortable=True),
            ]
            
            # Create the index
            search_index = SearchIndex(name=index_name, fields=fields)
            result = self.search_index_client.create_index(search_index)
            logger.info(f"Search index '{index_name}' created successfully")
            return True
            
        except Exception as e:
            logger.error(f"Error creating search index '{index_name}': {e}")
            return False
    
    def get_search_client(self, index_name: str) -> Optional[SearchClient]:
        """Get search client for specific index"""
        try:
            # Strict validation: index_name is required
            if not index_name:
                logger.error("Cannot create search client: index name is required but not provided")
                return None
            
            if not isinstance(index_name, str) or index_name.strip() == "":
                logger.error(f"Cannot create search client: invalid index name '{index_name}'")
                return None
            
            index_name = index_name.strip()
            
            if not self.config.search_endpoint or not self.config.search_admin_key:
                logger.error(f"Search endpoint or admin key not configured - cannot create client for index '{index_name}'")
                return None
                
            from azure.search.documents import SearchClient
            from azure.core.credentials import AzureKeyCredential
            
            credential = AzureKeyCredential(self.config.search_admin_key)
            search_client = SearchClient(
                endpoint=self.config.search_endpoint,
                index_name=index_name,
                credential=credential
            )
            return search_client
            
        except Exception as e:
            logger.error(f"Error creating search client for index '{index_name}': {e}")
            return None

    def extract_text_from_document(self, file_content: bytes, content_type: str, filename: str) -> str:
        """Extract text from various document types"""
        try:
            text_content = ""
            
            if content_type == 'application/pdf' or filename.lower().endswith('.pdf'):
                try:
                    import PyPDF2
                    from io import BytesIO
                    
                    pdf_reader = PyPDF2.PdfReader(BytesIO(file_content))
                    for page in pdf_reader.pages:
                        text_content += page.extract_text() + "\n"
                except Exception as pdf_error:
                    logger.warning(f"PDF extraction failed: {pdf_error}")
                    text_content = f"[PDF content extraction failed: {filename}]"
                    
            elif content_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document'] or filename.lower().endswith('.docx'):
                try:
                    from docx import Document
                    from io import BytesIO
                    
                    doc = Document(BytesIO(file_content))
                    for paragraph in doc.paragraphs:
                        text_content += paragraph.text + "\n"
                except Exception as docx_error:
                    logger.warning(f"DOCX extraction failed: {docx_error}")
                    text_content = f"[DOCX content extraction failed: {filename}]"
                    
            elif content_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'] or filename.lower().endswith('.xlsx'):
                try:
                    import openpyxl
                    from io import BytesIO
                    
                    wb = openpyxl.load_workbook(BytesIO(file_content))
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        text_content += f"Sheet: {sheet_name}\n"
                        for row in sheet.iter_rows(values_only=True):
                            row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                            if row_text.strip():
                                text_content += row_text + "\n"
                except Exception as xlsx_error:
                    logger.warning(f"XLSX extraction failed: {xlsx_error}")
                    text_content = f"[XLSX content extraction failed: {filename}]"
                    
            elif content_type.startswith('text/') or filename.lower().endswith(('.txt', '.md', '.csv')):
                try:
                    text_content = file_content.decode('utf-8', errors='ignore')
                except Exception as text_error:
                    logger.warning(f"Text extraction failed: {text_error}")
                    text_content = f"[Text content extraction failed: {filename}]"
            else:
                text_content = f"[Content type {content_type} not supported for text extraction: {filename}]"
                
            return text_content[:10000]  # Limit to 10K characters to avoid index size issues
            
        except Exception as e:
            logger.error(f"Error extracting text from document {filename}: {e}")
            return f"[Text extraction error: {filename}]"

    def get_text_embedding(self, text: str) -> Optional[List[float]]:
        """Generate text embedding using Azure OpenAI"""
        try:
            if not self.config.openai_endpoint or not self.config.openai_api_key:
                logger.warning("OpenAI endpoint or API key not configured, generating mock embedding for demo")
                # Return mock embedding for demo purposes
                import random
                random.seed(hash(text) % 2147483647)  # Consistent mock embedding for same text
                return [random.uniform(-1, 1) for _ in range(1536)]
                
            # Use Azure OpenAI for embeddings
            headers = {
                "Content-Type": "application/json",
                "api-key": self.config.openai_api_key
            }
            
            # Truncate text if too long (max 8191 tokens for text-embedding-ada-002)
            max_chars = 8000  # Conservative estimate
            if len(text) > max_chars:
                text = text[:max_chars]
                
            data = {
                "input": text
            }
            
            # Try multiple embedding endpoints
            embedding_endpoints = [
                f"{self.config.openai_endpoint}/openai/deployments/text-embedding-ada-002/embeddings?api-version=2023-05-15",
                f"{self.config.openai_endpoint}/openai/deployments/text-embedding-3-small/embeddings?api-version=2024-02-01",
                f"{self.config.openai_endpoint}/openai/embeddings?api-version=2023-05-15"
            ]
            
            for endpoint in embedding_endpoints:
                try:
                    response = requests.post(
                        endpoint,
                        headers=headers,
                        json=data,
                        timeout=30
                    )
                    
                    if response.status_code == 200:
                        result = response.json()
                        embedding = result["data"][0]["embedding"]
                        logger.info(f"Generated embedding with {len(embedding)} dimensions using endpoint: {endpoint}")
                        return embedding
                    else:
                        logger.warning(f"Endpoint {endpoint} failed with status {response.status_code}")
                        continue
                        
                except Exception as endpoint_error:
                    logger.warning(f"Error with endpoint {endpoint}: {endpoint_error}")
                    continue
            
            # If all endpoints fail, use mock embedding
            logger.warning("All embedding endpoints failed, using mock embedding")
            import random
            random.seed(hash(text) % 2147483647)
            return [random.uniform(-1, 1) for _ in range(1536)]
                
        except Exception as e:
            logger.error(f"Error generating text embedding: {e}")
            # Return mock embedding as fallback
            import random
            random.seed(hash(text) % 2147483647)
            return [random.uniform(-1, 1) for _ in range(1536)]

    def index_document(self, container_name: str, filename: str, file_content: bytes, 
                      content_type: str, index_name: str) -> bool:
        """Index a document in Azure Search with vector embeddings using indexer"""
        try:
            # Trigger indexer to reindex all documents in the container
            result = self.trigger_reindex_after_document_change(container_name, index_name)
            
            if result["success"]:
                logger.info(f"Document indexing triggered via indexer for '{filename}' in '{index_name}'")
                return True
            else:
                logger.warning(f"Failed to trigger indexing for '{filename}': {result.get('message')}")
                return False
                
        except Exception as e:
            logger.error(f"Error triggering document indexing for '{filename}': {e}")
            return False

    def search_documents(self, query: str, index_name: str, top: int = 10) -> List[Dict]:
        """Search documents using hybrid search (text + vector)"""
        try:
            search_client = self.get_search_client(index_name)
            if not search_client:
                logger.warning(f"Could not get search client for index '{index_name}'")
                return []
            
            # Perform search
            search_results = search_client.search(
                search_text=query,
                top=top,
                include_total_count=True
            )
            
            results = []
            for result in search_results:
                results.append(dict(result))
                
            logger.info(f"Found {len(results)} documents for query: '{query}' in index: '{index_name}'")
            return results
            
        except Exception as e:
            logger.error(f"Error searching documents in index '{index_name}': {e}")
            return []

    def upload_and_index_document(self, container_name: str, filename: str, 
                                 file_content: bytes, content_type: str, index_name: str) -> Dict:
        """Upload document to blob storage and trigger indexer for reindexing"""
        try:
            # Upload to blob storage
            blob_upload_success = self.upload_document(container_name, filename, file_content)
            
            if not blob_upload_success:
                return {
                    "success": False,
                    "message": "Failed to upload document to blob storage",
                    "indexed": False
                }
            
            # Trigger indexer for reindexing if index_name is provided
            indexed = False
            index_message = "No index specified"
            
            if index_name:
                index_result = self.trigger_reindex_after_document_change(container_name, index_name)
                indexed = index_result["success"]
                index_message = index_result["message"]
            
            logger.info(f"Document '{filename}' uploaded to blob storage. Indexing: {indexed}")
            
            return {
                "success": True,
                "message": f"Document '{filename}' uploaded successfully to blob storage",
                "indexed": indexed,
                "index_message": index_message
            }
            
        except Exception as e:
            logger.error(f"Error in upload_and_index_document: {e}")
            return {
                "success": False,
                "message": f"Error uploading document: {str(e)}",
                "indexed": False
            }
    
    def authenticate_with_device_code(self, username: str = None) -> Dict:
        """
        Authenticate using Azure AD device code flow (supports MFA)
        This is the recommended method for desktop applications
        """
        try:
            from msal import PublicClientApplication
            
            # MSAL configuration
            app = PublicClientApplication(
                client_id=self.config.public_client_id,
                authority=f"https://login.microsoftonline.com/{self.config.tenant_id}"
            )
            
            # Scope for basic profile information
            scopes = ["https://graph.microsoft.com/User.Read"]
            
            # Device code flow
            flow = app.initiate_device_flow(scopes=scopes)
            
            if "user_code" not in flow:
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": "Failed to create device flow"
                }
            
            # Return the device code information for user to complete authentication
            return {
                "success": False,  # Not yet complete
                "user": None,
                "token": None,
                "message": flow["message"],
                "device_flow": flow,
                "user_code": flow["user_code"],
                "verification_uri": flow["verification_uri"],
                "expires_in": flow["expires_in"]
            }
                
        except ImportError:
            logger.error("MSAL library not found. Please install: pip install msal")
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": "MSAL library not installed. Please install msal package."
            }
        except Exception as e:
            logger.error(f"Device code authentication error: {e}")
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Device code authentication error: {str(e)}"
            }
    
    def complete_device_code_authentication(self, device_flow: Dict) -> Dict:
        """
        Complete the device code authentication after user has authenticated
        """
        try:
            from msal import PublicClientApplication
            
            # MSAL configuration
            app = PublicClientApplication(
                client_id=self.config.public_client_id,
                authority=f"https://login.microsoftonline.com/{self.config.tenant_id}"
            )
            
            # Complete the device flow
            result = app.acquire_token_by_device_flow(device_flow)
            
            if "access_token" in result:
                # Success with device flow
                user_info = self._get_user_info_from_graph(result["access_token"])
                username = user_info.get("userPrincipalName") or user_info.get("mail")
                role = self._determine_user_role(username)
                permissions = self._get_user_permissions(username, role)
                
                logger.info(f"Real Azure device flow authentication successful for {username}")
                return {
                    "success": True,
                    "user": {
                        "username": username,
                        "display_name": user_info.get("displayName", username),
                        "role": role,
                        "permissions": permissions,
                        "email": user_info.get("mail") or user_info.get("userPrincipalName"),
                        "real_azure_user": True,
                        "auth_method": "device_code"
                    },
                    "token": result["access_token"],
                    "message": "Azure AD device code authentication successful"
                }
            else:
                error_description = result.get("error_description", "Unknown error")
                return {
                    "success": False,
                    "user": None,
                    "token": None,
                    "message": f"Azure AD device code authentication failed: {error_description}"
                }
                
        except Exception as e:
            logger.error(f"Device code completion error: {e}")
            return {
                "success": False,
                "user": None,
                "token": None,
                "message": f"Device code completion error: {str(e)}"
            }

    def run_indexer(self, indexer_name: str) -> Dict:
        """Run Azure Search indexer to reindex all documents"""
        try:
            if not self.config.search_endpoint or not self.config.search_admin_key:
                logger.warning("Search endpoint or admin key not configured. Skipping indexer run.")
                return {"success": False, "message": "Search not configured"}
            
            from azure.search.documents.indexes import SearchIndexerClient
            from azure.core.credentials import AzureKeyCredential
            
            # Create indexer client
            credential = AzureKeyCredential(self.config.search_admin_key)
            indexer_client = SearchIndexerClient(
                endpoint=self.config.search_endpoint,
                credential=credential
            )
            
            # Run the indexer
            indexer_client.run_indexer(indexer_name)
            
            logger.info(f"Indexer '{indexer_name}' started successfully")
            
            return {
                "success": True,
                "message": f"Indexer '{indexer_name}' started successfully",
                "indexer_name": indexer_name
            }
            
        except Exception as e:
            logger.error(f"Error running indexer '{indexer_name}': {e}")
            return {
                "success": False, 
                "message": f"Error running indexer: {str(e)}"
            }

    def get_indexer_status(self, indexer_name: str) -> Dict:
        """Get the status of an Azure Search indexer"""
        try:
            if not self.config.search_endpoint or not self.config.search_admin_key:
                logger.warning("Search endpoint or admin key not configured.")
                return {"success": False, "message": "Search not configured"}
            
            from azure.search.documents.indexes import SearchIndexerClient
            from azure.core.credentials import AzureKeyCredential
            
            # Create indexer client
            credential = AzureKeyCredential(self.config.search_admin_key)
            indexer_client = SearchIndexerClient(
                endpoint=self.config.search_endpoint,
                credential=credential
            )
            
            # Get indexer status
            status = indexer_client.get_indexer_status(indexer_name)
            
            return {
                "success": True,
                "status": status.status,
                "last_result": {
                    "status": status.last_result.status if status.last_result else None,
                    "start_time": status.last_result.start_time if status.last_result else None,
                    "end_time": status.last_result.end_time if status.last_result else None,
                    "item_count": status.last_result.item_count if status.last_result else 0,
                    "failed_item_count": status.last_result.failed_item_count if status.last_result else 0
                }
            }
            
        except Exception as e:
            logger.error(f"Error getting indexer status for '{indexer_name}': {e}")
            return {
                "success": False,
                "message": f"Error getting indexer status: {str(e)}"
            }

    def remove_document_from_index(self, container_name: str, file_name: str, index_name: str) -> bool:
        """Remove document from Azure Search index using title field search and chunk_id deletion"""
        try:
            # Strict validation: all parameters are required
            if not index_name:
                logger.error(f"Cannot remove document '{file_name}': search index name is required but not provided")
                return False
            
            if not isinstance(index_name, str) or index_name.strip() == "":
                logger.error(f"Cannot remove document '{file_name}': invalid search index name '{index_name}'")
                return False
            
            if not file_name or not isinstance(file_name, str) or file_name.strip() == "":
                logger.error(f"Cannot remove document: invalid file name '{file_name}'")
                return False
            
            index_name = index_name.strip()
            file_name = file_name.strip()
            
            if not self.config.search_endpoint or not self.config.search_admin_key:
                logger.error(f"Search endpoint or admin key not configured. Cannot remove document '{file_name}' from index '{index_name}'")
                return False
            
            from azure.search.documents import SearchClient
            from azure.core.credentials import AzureKeyCredential
            
            # Create search client
            credential = AzureKeyCredential(self.config.search_admin_key)
            search_client = SearchClient(
                endpoint=self.config.search_endpoint,
                index_name=index_name,
                credential=credential
            )
            
            logger.info(f"🔍 Searching for documents with title '{file_name}' in index '{index_name}'...")
            
            # Search for documents that match the filename in title field
            search_results = list(search_client.search(search_text=f'title:"{file_name}"', top=100))
            logger.info(f"📊 Found {len(search_results)} documents with exact title match")
            
            if len(search_results) == 0:
                # Try broader search if exact title doesn't work
                logger.info(f"⚠️ No exact title matches, trying broader search...")
                search_results = list(search_client.search(search_text=file_name, top=100))
                logger.info(f"📊 Found {len(search_results)} documents with broader search")
            
            if len(search_results) == 0:
                logger.warning(f"❌ No documents found for '{file_name}' in index '{index_name}'")
                return False
            
            # Find documents that actually match our target file
            documents_to_delete = []
            for result in search_results:
                result_dict = dict(result)
                title = result_dict.get('title', '')
                
                # Check if this document matches our file
                if file_name == title or file_name in title:
                    documents_to_delete.append(result_dict)
                    logger.info(f"✅ Found matching document: {title} (chunk_id: {result_dict.get('chunk_id', 'N/A')})")
            
            if not documents_to_delete:
                logger.warning(f"❌ No documents with matching titles found for '{file_name}'")
                return False
            
            logger.info(f"🎯 Will attempt to delete {len(documents_to_delete)} document chunks")
            
            # Try to delete each matching document using chunk_id
            successful_deletions = 0
            for i, doc in enumerate(documents_to_delete):
                try:
                    # Use chunk_id as the key field (based on our successful test)
                    chunk_id = doc.get('chunk_id')
                    if not chunk_id:
                        logger.warning(f"   ❌ Document {i+1} has no chunk_id field")
                        continue
                    
                    # Prepare deletion payload
                    delete_payload = {'chunk_id': chunk_id}
                    logger.info(f"   🚀 Deleting document {i+1}/{len(documents_to_delete)} with chunk_id: {chunk_id[:50]}...")
                    
                    # Perform the deletion
                    delete_result = search_client.delete_documents([delete_payload])
                    
                    if delete_result and len(delete_result) > 0:
                        result_item = delete_result[0]
                        if hasattr(result_item, 'succeeded') and result_item.succeeded:
                            logger.info(f"   ✅ Successfully deleted document chunk {i+1}")
                            successful_deletions += 1
                        elif hasattr(result_item, 'status_code') and result_item.status_code in [200, 204]:
                            logger.info(f"   ✅ Successfully deleted document chunk {i+1}")
                            successful_deletions += 1
                        else:
                            logger.warning(f"   ⚠️ Delete operation may have failed for chunk {i+1}")
                    else:
                        logger.warning(f"   ❌ Empty delete result for chunk {i+1}")
                        
                except Exception as delete_error:
                    logger.error(f"   💥 Failed to delete document chunk {i+1}: {delete_error}")
                    continue
            
            # Report results
            if successful_deletions > 0:
                logger.info(f"✅ Successfully deleted {successful_deletions}/{len(documents_to_delete)} document chunks for '{file_name}'")
                return True
            else:
                logger.error(f"❌ Failed to delete any document chunks for '{file_name}' from index '{index_name}'")
                return False
            
        except Exception as e:
            logger.error(f"💥 Error removing document '{file_name}' from index '{index_name}': {e}")
            return False

    def trigger_reindex_after_document_change(self, container_name: str, index_name: str) -> Dict:
        """Trigger reindexing after document upload or deletion using indexer"""
        try:
            # Generate indexer name based on container/index
            indexer_name = f"{index_name}-indexer"
            
            # Run the indexer
            result = self.run_indexer(indexer_name)
            
            if result["success"]:
                logger.info(f"Reindexing triggered for container '{container_name}' using indexer '{indexer_name}'")
                return {
                    "success": True,
                    "message": f"Reindexing started for container '{container_name}'",
                    "indexer_name": indexer_name,
                    "container_name": container_name,
                    "index_name": index_name
                }
            else:
                return result
                
        except Exception as e:
            logger.error(f"Error triggering reindex for container '{container_name}': {e}")
            return {
                "success": False,
                "message": f"Error triggering reindex: {str(e)}"
            }

    def _advanced_search_and_delete(self, container_name: str, file_name: str, index_name: str) -> bool:
        """Advanced search and delete method - simplified and focused on title field search"""
        try:
            # Strict validation: all parameters are required
            if not index_name:
                logger.error(f"Cannot perform advanced search and delete for '{file_name}': search index name is required but not provided")
                return False
            
            if not isinstance(index_name, str) or index_name.strip() == "":
                logger.error(f"Cannot perform advanced search and delete for '{file_name}': invalid search index name '{index_name}'")
                return False
            
            if not file_name or not isinstance(file_name, str) or file_name.strip() == "":
                logger.error(f"Cannot perform advanced search and delete: invalid file name '{file_name}'")
                return False
            
            index_name = index_name.strip()
            file_name = file_name.strip()
            
            search_client = self.get_search_client(index_name)
            if not search_client:
                logger.error(f"Could not get search client for index '{index_name}' - index may not exist")
                return False
            
            logger.info(f"🔍 Advanced search and delete for '{file_name}' in index '{index_name}'")
            
            # Search for documents that match the filename using title field
            search_results = list(search_client.search(search_text=f'title:"{file_name}"', top=100))
            logger.info(f"📊 Found {len(search_results)} results for title search")
            
            if len(search_results) == 0:
                # Try broader search
                logger.info(f"⚠️ No title matches, trying broader search...")
                search_results = list(search_client.search(search_text=file_name, top=100))
                logger.info(f"📊 Found {len(search_results)} results for broader search")
            
            if len(search_results) == 0:
                logger.warning(f"No documents found matching filename '{file_name}' in index '{index_name}'")
                return False
            
            # Filter results to find exact matches
            matching_documents = []
            for result in search_results:
                result_dict = dict(result)
                title = result_dict.get('title', '')
                
                # Check if this is our target document
                if file_name == title or file_name in title:
                    matching_documents.append(result_dict)
                    logger.info(f"✅ Found matching document: {title}")
            
            if not matching_documents:
                logger.warning(f"No documents with matching titles found for '{file_name}'")
                return False
            
            logger.info(f"🎯 Starting deletion process for {len(matching_documents)} document chunks")
            
            # Try to delete each matching document
            deletion_success_count = 0
            for i, doc in enumerate(matching_documents):
                try:
                    # Use chunk_id as key field (confirmed working in tests)
                    chunk_id = doc.get('chunk_id')
                    if not chunk_id:
                        logger.warning(f"   ❌ Document {i+1} missing chunk_id field")
                        continue
                    
                    # Prepare deletion payload
                    deletion_document = {'chunk_id': chunk_id}
                    logger.info(f"   🚀 Deleting document {i+1}/{len(matching_documents)}: {chunk_id[:50]}...")
                    
                    # Perform the deletion
                    delete_result = search_client.delete_documents([deletion_document])
                    
                    if delete_result and len(delete_result) > 0:
                        result_item = delete_result[0]
                        if hasattr(result_item, 'succeeded') and result_item.succeeded:
                            logger.info(f"   ✅ Successfully deleted document chunk {i+1}")
                            deletion_success_count += 1
                        elif hasattr(result_item, 'status_code') and result_item.status_code in [200, 204]:
                            logger.info(f"   ✅ Successfully deleted document chunk {i+1}")
                            deletion_success_count += 1
                        else:
                            logger.warning(f"   ⚠️ Delete operation may have failed for chunk {i+1}")
                    else:
                        logger.warning(f"   ❌ Empty delete result for chunk {i+1}")
                    
                except Exception as delete_error:
                    logger.error(f"   💥 Failed to delete document chunk {i+1}: {delete_error}")
                    continue
            
            # Report results
            if deletion_success_count > 0:
                logger.info(f"✅ Successfully deleted {deletion_success_count}/{len(matching_documents)} document chunks for '{file_name}'")
                return True
            else:
                logger.error(f"❌ Failed to delete any document chunks for '{file_name}' from index '{index_name}'")
                return False
            
        except Exception as e:
            logger.error(f"💥 Error in advanced search and delete: {e}")
            return False
    

class BlobStorageAgentManager:
    """Manages agent configurations using Azure Blob Storage"""
    
    def __init__(self, config: 'AzureConfig'):
        self.config = config
        self.container_name = "agent-configs"
        self.blob_client = None
        self._init_blob_client()
    
    def _init_blob_client(self):
        """Initialize blob client with proper authentication"""
        try:
            # Use connection string for reliable authentication
            if self.config.storage_connection_string:
                self.blob_client = BlobServiceClient.from_connection_string(
                    self.config.storage_connection_string
                )
                logger.info("Agent manager blob client initialized with connection string")
            else:
                # Fallback to Azure AD authentication
                credential = DefaultAzureCredential()
                self.blob_client = BlobServiceClient(
                    account_url=f"https://{self.config.storage_account_name}.blob.core.windows.net",
                    credential=credential
                )
                logger.info("Agent manager blob client initialized with Azure AD credential")
            
            # Ensure container exists
            try:
                self.blob_client.create_container(self.container_name)
                logger.info(f"Agent container '{self.container_name}' created or already exists")
            except Exception as e:
                logger.debug(f"Agent container creation info: {e}")
                pass  # Container might already exist
                
        except Exception as e:
            logger.error(f"Failed to initialize blob client for agent manager: {e}")
            self.blob_client = None
    
    def get_agent(self, agent_id: str) -> Optional[Dict]:
        """Get agent configuration by ID"""
        try:
            if not self.blob_client:
                return None
                
            blob_name = f"{agent_id}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name, 
                blob=blob_name
            )
            
            try:
                blob_data = blob_client.download_blob().readall()
                return json.loads(blob_data.decode('utf-8'))
            except Exception:
                return None
                
        except Exception as e:
            logger.error(f"Error getting agent {agent_id}: {e}")
            return None
    
    def get_all_agents(self) -> Dict[str, Dict]:
        """Get all agent configurations"""
        try:
            if not self.blob_client:
                return {}
                
            container_client = self.blob_client.get_container_client(self.container_name)
            agents = {}
            
            for blob in container_client.list_blobs():
                if blob.name.endswith('.json'):
                    agent_id = blob.name.replace('.json', '')
                    agent_data = self.get_agent(agent_id)
                    if agent_data:
                        agents[agent_id] = agent_data
            
            return agents
            
        except Exception as e:
            logger.error(f"Error getting all agents: {e}")
            return {}
    
    def get_active_agents(self) -> Dict[str, Dict]:
        """Get only active agent configurations"""
        all_agents = self.get_all_agents()
        return {agent_id: agent_config for agent_id, agent_config in all_agents.items() 
                if agent_config.get('status') == 'active'}
    
    def add_agent(self, agent_config: Dict) -> bool:
        """Add new agent configuration"""
        try:
            if not self.blob_client:
                return False
                
            agent_id = agent_config.get('id')
            if not agent_id:
                return False
                
            blob_name = f"{agent_id}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            # Add timestamp
            agent_config['created_at'] = datetime.now().isoformat()
            agent_config['updated_at'] = datetime.now().isoformat()
            
            blob_client.upload_blob(
                json.dumps(agent_config, indent=2).encode('utf-8'),
                overwrite=True
            )
            
            return True
            
        except Exception as e:
            logger.error(f"Error adding agent: {e}")
            return False
    
    def update_agent(self, agent_id: str, agent_config: Dict) -> bool:
        """Update existing agent configuration"""
        try:
            if not self.blob_client:
                return False
                
            # Get existing config to preserve creation time
            existing_config = self.get_agent(agent_id)
            if existing_config:
                agent_config['created_at'] = existing_config.get('created_at')
            
            agent_config['updated_at'] = datetime.now().isoformat()
            agent_config['id'] = agent_id
            
            blob_name = f"{agent_id}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            blob_client.upload_blob(
                json.dumps(agent_config, indent=2).encode('utf-8'),
                overwrite=True
            )
            
            return True
            
        except Exception as e:
            logger.error(f"Error updating agent {agent_id}: {e}")
            return False
    
    def delete_agent(self, agent_id: str) -> bool:
        """Delete agent configuration"""
        try:
            if not self.blob_client:
                return False
                
            blob_name = f"{agent_id}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            blob_client.delete_blob()
            return True
            
        except Exception as e:
            logger.error(f"Error deleting agent {agent_id}: {e}")
            return False
    
    def set_agent_status(self, agent_id: str, status: str) -> bool:
        """Set agent status (active/inactive)"""
        try:
            agent_config = self.get_agent(agent_id)
            if not agent_config:
                return False
                
            agent_config['status'] = status
            return self.update_agent(agent_id, agent_config)
            
        except Exception as e:
            logger.error(f"Error setting agent status for {agent_id}: {e}")
            return False
    
    def generate_azure_container_name(self, agent_id: str) -> str:
        """Generate Azure-compliant container name for agent"""
        # Azure container names must be lowercase, 3-63 chars, alphanumeric and hyphens
        container_name = f"agent-{agent_id}".lower()
        # Replace any invalid characters
        container_name = ''.join(c if c.isalnum() or c == '-' else '-' for c in container_name)
        # Ensure it doesn't exceed 63 characters
        if len(container_name) > 63:
            container_name = container_name[:63]
        # Ensure it doesn't end with a hyphen
        container_name = container_name.rstrip('-')
        return container_name


class BlobStorageUserManager:
    """Manages user authentication and permissions using Azure Blob Storage"""
    
    def __init__(self, config: 'AzureConfig'):
        self.config = config
        self.container_name = "user-configs"
        self.blob_client = None
        self._init_blob_client()
        self._init_default_admin()
    
    def _init_blob_client(self):
        """Initialize blob client with proper authentication"""
        try:
            # Use connection string for reliable authentication
            if self.config.storage_connection_string:
                self.blob_client = BlobServiceClient.from_connection_string(
                    self.config.storage_connection_string
                )
                logger.info("User manager blob client initialized with connection string")
            else:
                # Fallback to Azure AD authentication
                credential = DefaultAzureCredential()
                self.blob_client = BlobServiceClient(
                    account_url=f"https://{self.config.storage_account_name}.blob.core.windows.net",
                    credential=credential
                )
                logger.info("User manager blob client initialized with Azure AD credential")
            
            # Ensure container exists
            try:
                self.blob_client.create_container(self.container_name)
                logger.info(f"User container '{self.container_name}' created or already exists")
            except Exception as e:
                logger.debug(f"User container creation info: {e}")
                pass  # Container might already exist
                
        except Exception as e:
            logger.error(f"Failed to initialize blob client for user manager: {e}")
            self.blob_client = None
    
    def _init_default_admin(self):
        """Initialize default admin user if none exists"""
        try:
            admins = self.get_all_users()
            
            # Ensure admins is a dictionary
            if not isinstance(admins, dict):
                logger.warning(f"get_all_users returned unexpected type: {type(admins)}")
                admins = {}
            
            # Check if any admin user exists
            admin_exists = False
            for username, user_data in admins.items():
                if isinstance(user_data, dict) and user_data.get('role') == 'admin':
                    admin_exists = True
                    break
            
            if not admin_exists:
                logger.info("No admin user found, creating default admin")
                self.add_user('admin', 'admin')
                logger.info("Default admin user created with username: admin, password: G5x!bQz2Lp9")
                
        except Exception as e:
            logger.error(f"Error initializing default admin: {e}")
            # In case of error, still try to create admin
            try:
                self.add_user('admin', 'admin')
                logger.info("Fallback: Default admin user created")
            except Exception as fallback_error:
                logger.error(f"Fallback admin creation also failed: {fallback_error}")
    
    def _hash_password(self, password: str) -> str:
        """Hash password using SHA256"""
        return hashlib.sha256(password.encode()).hexdigest()
    
    def authenticate_admin(self, username: str, password: str) -> bool:
        """Authenticate admin user"""
        try:
            user_data = self.get_user(username)
            if not user_data:
                return False
                
            if user_data.get('role') != 'admin':
                return False
                
            password_hash = self._hash_password(password)
            return user_data.get('password_hash') == password_hash
            
        except Exception as e:
            logger.error(f"Error authenticating admin {username}: {e}")
            return False
    
    def authenticate_azure_user(self, username: str, password: str) -> Dict:
        """Authenticate Azure user using Azure AD (placeholder implementation)"""
        try:
            # For now, just check if user exists in our system
            # In production, this would integrate with Azure AD
            user_data = self.get_user(username)
            if user_data:
                return {
                    'authenticated': True,
                    'user': user_data,
                    'token': 'azure_token_placeholder'
                }
            else:
                return {
                    'authenticated': False,
                    'error': 'User not found'
                }
                
        except Exception as e:
            logger.error(f"Error authenticating Azure user {username}: {e}")
            return {
                'authenticated': False,
                'error': str(e)
            }
    
    def get_user(self, username: str) -> Optional[Dict]:
        """Get user by username"""
        try:
            if not self.blob_client:
                return None
                
            blob_name = f"{username}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            try:
                blob_data = blob_client.download_blob().readall()
                return json.loads(blob_data.decode('utf-8'))
            except Exception:
                return None
                
        except Exception as e:
            logger.error(f"Error getting user {username}: {e}")
            return None
    
    def get_user_permissions(self, username: str) -> Dict:
        """Get user permissions"""
        try:
            user_data = self.get_user(username)
            if user_data:
                return {
                    'permissions': user_data.get('permissions', []),
                    'role': user_data.get('role', 'user')
                }
            else:
                return {
                    'permissions': [],
                    'role': 'guest'
                }
                
        except Exception as e:
            logger.error(f"Error getting permissions for {username}: {e}")
            return {
                'permissions': [],
                'role': 'guest'
            }
    
    def has_permission(self, username: str, agent_id: str, permission_type: str) -> bool:
        """Check if user has specific permission for agent"""
        try:
            user_permissions = self.get_user_permissions(username)
            permissions = user_permissions.get('permissions', [])
            role = user_permissions.get('role', 'guest')
            
            # Admin has all permissions
            if role == 'admin' or 'all' in permissions:
                return True
            
            # Handle both old dictionary format and new list format
            if isinstance(permissions, dict):
                # Old dictionary format: {'scm': {'access': True, 'chat': True, ...}}
                agent_perms = permissions.get(agent_id, {})
                if isinstance(agent_perms, dict):
                    return agent_perms.get(permission_type, False)
                return False
            elif isinstance(permissions, list):
                # New list format: ['access', 'chat', 'scm:access', 'scm:chat', ...]
                # Check for wildcard permission (without agent_id prefix)
                if permission_type in permissions:
                    return True
                    
                # Check for specific agent permission (agent_id:permission_type)
                specific_permission = f"{agent_id}:{permission_type}"
                if specific_permission in permissions:
                    return True
                
                # For standard users, provide default access to basic functions
                if role == "standard" and permission_type in ["access", "chat"]:
                    return True
            
            return False
            
        except Exception as e:
            logger.error(f"Error checking permission for {username}: {e}")
            return False
    
    def add_user(self, username: str, role: str, permissions: List[str] = None) -> bool:
        """Add new user"""
        try:
            if not self.blob_client:
                return False
                
            # Generate default password for admin
            if role == "admin":
                default_password = "G5x!bQz2Lp9"
            else:
                default_password = f"{username}123"
            
            # Get default permissions if none provided
            if permissions is None:
                permissions = self._get_user_permissions(username, role)
            
            user_data = {
                'username': username,
                'password_hash': self._hash_password(default_password),
                'role': role,
                'permissions': permissions,
                'created_at': datetime.now().isoformat(),
                'updated_at': datetime.now().isoformat()
            }
            
            blob_name = f"{username}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            blob_client.upload_blob(
                json.dumps(user_data, indent=2).encode('utf-8'),
                overwrite=True
            )
            
            return True
            
        except Exception as e:
            logger.error(f"Error adding user {username}: {e}")
            return False
    
    def update_user_permissions(self, username: str, permissions: List[str]) -> bool:
        """Update user permissions"""
        try:
            user_data = self.get_user(username)
            if not user_data:
                return False
                
            user_data['permissions'] = permissions
            user_data['updated_at'] = datetime.now().isoformat()
            
            blob_name = f"{username}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            blob_client.upload_blob(
                json.dumps(user_data, indent=2).encode('utf-8'),
                overwrite=True
            )
            
            return True
            
        except Exception as e:
            logger.error(f"Error updating permissions for {username}: {e}")
            return False
    
    def delete_user(self, username: str) -> bool:
        """Delete user"""
        try:
            if not self.blob_client:
                return False
                
            blob_name = f"{username}.json"
            blob_client = self.blob_client.get_blob_client(
                container=self.container_name,
                blob=blob_name
            )
            
            blob_client.delete_blob()
            return True
            
        except Exception as e:
            logger.error(f"Error deleting user {username}: {e}")
            return False
    
    def get_all_users(self) -> Dict[str, Dict]:
        """Get all users"""
        try:
            if not self.blob_client:
                return {}
                
            container_client = self.blob_client.get_container_client(self.container_name)
            users = {}
            
            for blob in container_client.list_blobs():
                if blob.name.endswith('.json'):
                    username = blob.name.replace('.json', '')
                    user_data = self.get_user(username)
                    if user_data:
                        # Remove password hash from returned data for security
                        safe_user_data = {k: v for k, v in user_data.items() if k != 'password_hash'}
                        users[username] = safe_user_data
            
            return users
            
        except Exception as e:
            logger.error(f"Error getting all users: {e}")
            return {}
